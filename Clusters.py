#FOR MONGO AND EXCEL
import json
import requests
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

#####################################################################
#				ONLY PARAMETERS THAT CAN BE MODIFIED				#
# 					DO NOT MODIFY ANYTHING ELSE 					#
#####################################################################

Token = ''
aid = 155958

#####################################################################
#####################################################################



headers = {
    'Authorization': 'Bearer {}'.format(Token),
    'Content-Type': 'application/json'
}

startTime = time.time()

s = requests.Session()
URL = 'https://api.thousandeyes.com/v6/agents.json?aid=%s&agentTypes=ENTERPRISE_CLUSTER,ENTERPRISE' % aid
agents_info = s.get(URL, headers=headers).json()
agents_info = agents_info['agents']


#EXCEL
wb = load_workbook('Clusters-info.xlsx')
ws = wb.active

created = 0
edited = 0
errors = [0]
warnings = [0]

for i in range(2,ws.max_row+1):
    print("Script in progress"+ "#"*(i-1))
    agent_id=0
    cluster_id=0
    #Does the agent and cluster already exists? If so, save their IDs
    for agent in agents_info:
        if (agent["agentName"] ==  ws.cell(row=i,column=2).value or agent["agentName"] ==  ws.cell(row=i,column=1).value) and agent["agentType"] == "Enterprise Cluster":
            cluster_id = agent["agentId"]

        elif agent["agentName"] ==  ws.cell(row=i,column=1).value:
            agent_id = agent["agentId"]
            agents_info.remove(agent)


    #If the agent and the cluster were identified, we simply add the agent to the cluster.
    if agent_id != 0 and cluster_id != 0:
        edited +=1 
        URL_add = 'https://api.thousandeyes.com/v6/agents/%s/add-to-cluster.json?aid=%s' % (cluster_id, aid)
        payload = json.dumps([agent_id])
        result= s.post(URL_add, headers=headers,data=payload)
        value = "%s agent was added to %s cluster" % (ws.cell(row=i,column=1).value, ws.cell(row=i,column=2).value)
        ws.cell(row=i, column=3, value=str(value)).alignment = Alignment(horizontal="center",wrap_text=True)
        ws.cell(row=i, column=3 ).font = Font(color='000000')



    #If the agent was identified but the cluster wasn't this will create the cluster with the name assigned on the excel file
    elif agent_id != 0 and cluster_id == 0:
        created +=1
        URL_add = 'https://api.thousandeyes.com/v6/agents/%s/add-to-cluster.json?aid=%s' % (agent_id, aid)
        payload = json.dumps([])
        result= s.post(URL_add, headers=headers, data=payload).json()
        
        result = result["agents"]
        new_cluster = {
            "agentId":result[0]["agentId"],
            "agentName":ws.cell(row=i,column=2).value,
            "agentType":result[0]["agentType"]
        }
        agents_info.append(new_cluster)
        URL_edit= 'https://api.thousandeyes.com/v6/agents/%s/update.json?aid=%s' % (new_cluster["agentId"], aid)
        payload = json.dumps({"agentName":ws.cell(row=i,column=2).value})
        result= s.post(URL_edit, headers=headers, data=payload)
        
        if ws.cell(row=i,column=2).value != None:
            if result.status_code == 200:
                value = "%s cluster was created and %s agent was added" % (ws.cell(row=i,column=2).value, ws.cell(row=i,column=1).value)
                ws.cell(row=i, column=3, value=str(value)).alignment = Alignment(horizontal="center",wrap_text=True)
                ws.cell(row=i, column=3 ).font = Font(color='000000')
            else:
                value = "%s cluster's name is being used by another agent in your organization. The name of the cluster assigned will be %s" % (ws.cell(row=i,column=1).value, ws.cell(row=i,column=1).value)
                ws.cell(row=i, column=3, value=str(value)).alignment = Alignment(horizontal="center",wrap_text=True)
                ws.cell(row=i, column=3 ).font = Font(color='ff9933')
                warnings[0] += 1
                warnings.append(ws.cell(row=i,column=1).value)
        else: 
            new_cluster["agentName"] = ws.cell(row=i,column=1).value
            value = "%s cluster was created and %s agent was added" % (ws.cell(row=i,column=1).value, ws.cell(row=i,column=1).value)
            ws.cell(row=i, column=3, value=str(value)).alignment = Alignment(horizontal="center",wrap_text=True)
            ws.cell(row=i, column=3 ).font = Font(color='000000')


    #If the agent was not found, no cluster can be created or edited so it will just write some logs on the excel file
    else:
        value = "%s agent does not exist so cannot be added to a cluster" % (ws.cell(row=i,column=1).value)
        ws.cell(row=i, column=3, value=str(value)).alignment = Alignment(horizontal="center",wrap_text=True)  
        ws.cell(row=i, column=3 ).font = Font(color='ff0400')
        errors[0] += 1
        errors.append(ws.cell(row=i,column=1).value)

ws.column_dimensions[get_column_letter(3)].width = 70
wb.save("Clusters-info.xlsx")

print("=============== S U M M A R Y ==================")
print("This script has finished, time elapsed:"+ str(time.time() - startTime)[:7])
print("Clusters created:", created)
print("Clusters edited:", edited)
print("Agents added:", created+edited)
print("Errors detected: %d on this agents: %s" % (errors[0], str(errors[1:])))
print("Warnings detected %d on this agents: %s" % (warnings[0], str(warnings[1:])))







